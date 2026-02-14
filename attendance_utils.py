from calendar import monthrange
from datetime import datetime
from datetime import time as dt_time
from io import BytesIO
from typing import List

import pandas as pd

# -------------------- Core Data Extraction & Cleanup Helpers -------------------- #

def parse_sheet_name(sheet_name: str) -> List[str]:
    """Given a sheet name like '1.2.3', return a list of staff IDs (strings)."""
    return [s.strip() for s in str(sheet_name).split('.') if s and s.strip()]


def _normalise_time(val) -> str:
    """Convert a cell value to a normalised HH:MM string.

    Handles:
      - None / NaN  → ''
      - datetime.time → '08:30'
      - datetime.datetime → '08:30' (time part only)
      - float (Excel serial fraction e.g. 0.354) → '08:30'
      - str already in HH:MM → kept as-is
      - str with junk → stripped and returned
    """
    if val is None:
        return ''
    if isinstance(val, float) and pd.isna(val):
        return ''
    if isinstance(val, dt_time):
        return val.strftime('%H:%M')
    if isinstance(val, datetime):
        return val.strftime('%H:%M')
    if isinstance(val, (int, float)):
        # Excel stores times as fractional days (0.354166… = 08:30)
        try:
            total_seconds = int(round(float(val) * 86400))
            hours, remainder = divmod(abs(total_seconds), 3600)
            minutes = remainder // 60
            return f'{hours:02d}:{minutes:02d}'
        except (ValueError, OverflowError):
            return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'none', 'nat', ''):
        return ''
    return s


def convert_date(date_str: str, year: int, month: int) -> str:
    """Convert a date string in the format '01 SAT' to ISO 'YYYY-MM-DD'.

    Also accepts bare integers ('1', '15') and ISO dates ('2025-10-04').
    """
    s = str(date_str).strip()
    if not s:
        raise ValueError("Date string is empty")

    # Already ISO date?
    if '-' in s and len(s) >= 8:
        try:
            dt = datetime.fromisoformat(s[:10])
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass

    parts = s.split()
    if not parts:
        raise ValueError(f"Date string '{date_str}' is empty or invalid")
    try:
        day = int(parts[0])
    except (ValueError, TypeError) as e:
        raise ValueError(f"Cannot parse day from '{date_str}'") from e
    try:
        dt = datetime(year, month, day)
    except (ValueError, TypeError) as e:
        raise ValueError(f"Invalid date: year={year} month={month} day={day}") from e
    return dt.strftime('%Y-%m-%d')


# Column stride per staff block.  Each staff occupies 15 columns in the raw
# ZKTeco export sheet: Date(shared) + [blank, On, blank, Off, <11 metric cols>].
# Staff 1 starts at column 1 (On=1, Off=3),
# Staff 2 at column 16 (On=16, Off=18),
# Staff 3 at column 31 (On=31, Off=33).
_STAFF_OFFSETS = [
    (1, 3),    # Staff 1: On-duty col, Off-duty col
    (16, 18),  # Staff 2
    (31, 33),  # Staff 3
]


def process_sheet(sheet_name: str, df: pd.DataFrame, year: int, month: int):
    """Process a single worksheet according to fixed column layout.

    Column 0 is the shared date column.  Each staff block is at a fixed offset
    defined by _STAFF_OFFSETS.
    """
    records = []
    staff_ids = parse_sheet_name(sheet_name)
    num_cols = len(df.columns)

    for _idx, row in df.iterrows():
        date_raw = str(row.iloc[0]).strip()
        try:
            iso_date = convert_date(date_raw, year, month)
        except (ValueError, TypeError):
            continue  # skip non-date rows (totals, blanks, etc.)

        for i, staff_id in enumerate(staff_ids[:3]):
            if i >= len(_STAFF_OFFSETS):
                break
            on_col, off_col = _STAFF_OFFSETS[i]
            on_duty = _normalise_time(row.iloc[on_col] if on_col < num_cols else None)
            off_duty = _normalise_time(row.iloc[off_col] if off_col < num_cols else None)

            records.append({
                'ID': staff_id,
                'Date': iso_date,
                'OnDuty': on_duty,
                'OffDuty': off_duty,
                'Name': '',
                'Department': 'Company',
                'Late time(Min)': 0,
                'Leave early(Min)': 0,
                'Absence(Min)': 0,
                'Total(Min)': 0,
                'Note': ''
            })
    return records


def combine_all_sheets(excel_file, year: int, month: int) -> pd.DataFrame:
    """Read all valid sheets from the workbook and combine into a single DataFrame."""
    # Determine engine based on filename / stream name
    filename = getattr(excel_file, 'filename', None) or getattr(excel_file, 'name', '') or ''
    lower = filename.lower()
    engine = None
    if lower.endswith('.xls') and not lower.endswith('.xlsx'):
        engine = 'xlrd'
    elif lower.endswith('.xlsx'):
        engine = 'openpyxl'

    # Ensure stream is at the start
    if hasattr(excel_file, 'seek'):
        excel_file.seek(0)

    # Build ExcelFile once — all subsequent reads will use this cached object
    # instead of re-reading the stream (which would fail on non-seekable streams).
    kwargs = {'engine': engine} if engine else {}
    xls = pd.ExcelFile(excel_file, **kwargs)

    all_records: list = []
    for sheet_name in xls.sheet_names:
        staff_ids = parse_sheet_name(sheet_name)
        if not staff_ids or not all(s.isdigit() for s in staff_ids):
            continue
        # Read from the cached ExcelFile object (no stream reuse issue)
        df = xls.parse(sheet_name, header=None, skiprows=11)
        all_records.extend(process_sheet(sheet_name, df, year, month))

    final_df = pd.DataFrame(all_records)
    if final_df.empty:
        return final_df

    # Filter out rows where duty columns contain 'ABSENT'
    mask_absent = (
        final_df['OnDuty'].astype(str).str.contains('ABSENT', case=False, na=False) |
        final_df['OffDuty'].astype(str).str.contains('ABSENT', case=False, na=False)
    )
    final_df = final_df[~mask_absent].copy()

    # Also drop rows where both On and Off duty are blank (no clock data)
    mask_empty = (final_df['OnDuty'] == '') & (final_df['OffDuty'] == '')
    final_df = final_df[~mask_empty].copy()

    final_df['ID'] = final_df['ID'].astype(int)
    final_df = final_df.sort_values(['ID', 'Date']).reset_index(drop=True)

    final_df.rename(columns={
        'OnDuty': 'First time zone On-duty',
        'OffDuty': 'First time zone Off-duty',
    }, inplace=True)
    final_df['Second time zone On-duty'] = ''
    final_df['Second time zone Off-duty'] = ''

    # Ensure numeric metric columns are zero
    for col in ['Late time(Min)', 'Leave early(Min)', 'Absence(Min)', 'Total(Min)']:
        if col in final_df.columns:
            final_df[col] = 0

    cols = [
        'ID', 'Name', 'Department', 'Date',
        'First time zone On-duty', 'First time zone Off-duty',
        'Second time zone On-duty', 'Second time zone Off-duty',
        'Late time(Min)', 'Leave early(Min)',
        'Absence(Min)', 'Total(Min)', 'Note',
    ]
    final_df = final_df[cols]
    return final_df


def export_with_custom_header_to_bytes(final_df: pd.DataFrame, start_date_str: str, end_date_str: str) -> BytesIO:
    """Build the Exception Statistic Report workbook with a merged header."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Exception Stat.'

    # Row 1: Report title
    ws['A1'] = 'Exception Statistic Report'
    # Row 2: Date range
    ws['A2'] = 'Stat.Date:'
    ws['B2'] = f'{start_date_str} ~ {end_date_str}'
    # Row 3: Column headers
    ws['A3'] = 'ID'
    ws['B3'] = 'Name'
    ws['C3'] = 'Department'
    ws['D3'] = 'Date'
    ws['E3'] = 'First time zone'
    ws.merge_cells('E3:F3')
    ws['G3'] = 'Second time zone'
    ws.merge_cells('G3:H3')
    ws['I3'] = 'Late time(Min)'
    ws['J3'] = 'Leave early(Min)'
    ws['K3'] = 'Absence(Min)'
    ws['L3'] = 'Total(Min)'
    ws['M3'] = 'Note'
    # Row 4: Sub-headers for time-zone columns
    ws['E4'] = 'On-duty'
    ws['F4'] = 'Off-duty'
    ws['G4'] = 'On-duty'
    ws['H4'] = 'Off-duty'

    # Row 5+: Data
    for r_idx, row in enumerate(final_df.values, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell_val = val
            # Clean any residual NaN/None that slipped through
            if isinstance(cell_val, float) and pd.isna(cell_val):
                cell_val = ''
            elif cell_val is None:
                cell_val = ''
            ws.cell(row=r_idx, column=c_idx, value=cell_val)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def compute_date_range(year: int, month: int):
    """Return (start_date, end_date) strings for the given month."""
    days = monthrange(year, month)[1]
    return f"{year}-{month:02d}-01", f"{year}-{month:02d}-{days}"